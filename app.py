# -*- coding: utf-8 -*-
"""
HI-DRIVE: Sistema Avanzado de Gestión de Inventario con IA
Versión 3.0 - Optimizada (Caché RAM + Carga Asíncrona)
"""
import streamlit as st
from PIL import Image
import pandas as pd
import json
from datetime import datetime, timedelta, timezone
import numpy as np
import io  

# --- Importaciones de utilidades y modelos ---
try:
    from firebase_utils import FirebaseManager
    from gemini_utils import GeminiUtils
    from barcode_manager import BarcodeManager
    from twilio.rest import Client
    IS_TWILIO_AVAILABLE = True
except ImportError as e:
    st.error(f"Error de importación: {e}. Asegúrate de que todas las dependencias estén instaladas.")
    st.stop()


# --- CONFIGURACIÓN DE PÁGINA Y ESTILOS ---
st.set_page_config(
    page_title="Rapi Tienda Acuarela | Sistema SAVA",
    page_icon="https://github.com/GIUSEPPESAN21/LOGO-SAVA/blob/main/LOGO%20COLIBRI.png?raw=true",
    layout="wide"
)

# --- INYECCIÓN DE CSS ---
@st.cache_data
def load_css():
    try:
        with open("style.css") as f:
            st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)
    except FileNotFoundError:
        st.warning("Archivo style.css no encontrado. Se usarán estilos por defecto.")

load_css()

# --- INICIALIZACIÓN DE SERVICIOS ---
@st.cache_resource
def initialize_services():
    try:
        firebase_handler = FirebaseManager()
        barcode_handler = BarcodeManager(firebase_handler)
        gemini_handler = GeminiUtils()

        twilio_client = None
        if IS_TWILIO_AVAILABLE and all(k in st.secrets for k in ["TWILIO_ACCOUNT_SID", "TWILIO_AUTH_TOKEN", "TWILIO_WHATSAPP_FROM_NUMBER", "DESTINATION_WHATSAPP_NUMBER"]):
            try:
                twilio_client = Client(st.secrets["TWILIO_ACCOUNT_SID"], st.secrets["TWILIO_AUTH_TOKEN"])
            except Exception as twilio_e:
                st.warning(f"No se pudo inicializar Twilio: {twilio_e}.")
                twilio_client = None 
        else:
             pass

        return firebase_handler, gemini_handler, twilio_client, barcode_handler
    except Exception as e:
        st.error(f"**Error Crítico de Inicialización:** {e}")
        return None, None, None, None

firebase, gemini, twilio_client, barcode_manager = initialize_services()

if not all([firebase, gemini, barcode_manager]):
    st.error("Error al inicializar servicios esenciales. La aplicación no puede continuar.")
    st.stop()

# --- FUNCIONES CON CACHÉ EN RAM PARA ACELERAR EL RENDIMIENTO ---
@st.cache_data(ttl=300)
def get_cached_inventory():
    try: return firebase.get_all_inventory_items()
    except Exception: return []

@st.cache_data(ttl=300)
def get_cached_orders():
    try: return firebase.get_orders(status=None)
    except Exception: return []

@st.cache_data(ttl=300)
def get_cached_suppliers():
    try: return firebase.get_all_suppliers()
    except Exception: return []

# --- Funciones de Estado de Sesión ---
def init_session_state():
    defaults = {
        'page': "🏠 Inicio", 'order_items': [],
        'editing_item_id': None, 'scanned_item_data': None,
        'usb_scan_result': None, 'usb_sale_items': [],
        'add_sku_input': "", 
        'new_item_name': "", 'new_item_qty': 1, 
        'new_item_purchase': 0.0, 'new_item_sale': 0.0, 'new_item_alert': 0,
        'new_item_supplier': "",
        'should_clear_inventory_form': False
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

init_session_state()

# --- DIÁLOGOS DE INTERFAZ ---
@st.dialog("⚠️ Confirmar Eliminación")
def show_delete_confirmation(item_id, item_name):
    st.write(f"¿Estás seguro que deseas eliminar permanentemente el producto **{item_name}**?")
    st.warning("Esta acción borrará el inventario y el historial asociado. No se puede deshacer.")
    
    col1, col2 = st.columns(2)
    if col1.button("🚨 SÍ, ELIMINAR", type="primary", use_container_width=True):
        try:
            with st.spinner("Eliminando..."):
                firebase.delete_inventory_item(item_id)
                get_cached_inventory.clear() # <- Limpia caché
            st.success("¡Producto eliminado correctamente!")
            st.rerun()
        except Exception as e:
            st.error(f"Error al eliminar: {e}")
            
    if col2.button("Cancelar", use_container_width=True):
        st.rerun()

# --- LÓGICA DE NOTIFICACIONES ---
def send_whatsapp_alert(message):
    if not twilio_client: return
    try:
        from_number = st.secrets["TWILIO_WHATSAPP_FROM_NUMBER"]
        to_number = st.secrets["DESTINATION_WHATSAPP_NUMBER"]
        twilio_client.messages.create(from_=f'whatsapp:{from_number}', body=message, to=f'whatsapp:{to_number}')
        st.toast("¡Alerta de WhatsApp enviada!", icon="📲")
    except Exception as e:
        st.error(f"Error al enviar alerta de Twilio: {e}", icon="🚨")

# --- FUNCIONES CALLBACK & HELPERS ---
def set_clear_form_flag():
    st.session_state.should_clear_inventory_form = True

def save_new_item_callback(supplier_map, current_sku):
    name = st.session_state.get('new_item_name')
    quantity = st.session_state.get('new_item_qty')
    purchase_price = st.session_state.get('new_item_purchase')
    sale_price = st.session_state.get('new_item_sale')
    min_stock_alert = st.session_state.get('new_item_alert')
    selected_supplier_name = st.session_state.get('new_item_supplier')

    if not name:
        st.toast("⚠️ El nombre del artículo es obligatorio.", icon="⚠️")
        return

    supplier_id = supplier_map.get(selected_supplier_name)
    
    data = {
        "name": name,
        "quantity": int(quantity),
        "purchase_price": float(purchase_price),
        "sale_price": float(sale_price),
        "min_stock_alert": int(min_stock_alert),
        "supplier_id": supplier_id,
        "supplier_name": selected_supplier_name if supplier_id else "",
        "updated_at": datetime.now(timezone.utc).isoformat()
    }

    try:
        firebase.save_inventory_item(data, current_sku, is_new=True)
        get_cached_inventory.clear() # <- Limpia caché
        st.toast(f"✅ ¡Producto '{name}' guardado correctamente!", icon="✅")
        st.session_state.should_clear_inventory_form = True
    except Exception as add_e:
        st.toast(f"❌ Error al guardar: {add_e}", icon="❌")


# --- NAVEGACIÓN PRINCIPAL (SIDEBAR) ---
col1, col2, col3 = st.sidebar.columns([1,6,1])
with col2:
    st.image("https://github.com/GIUSEPPESAN21/LOGO-SAVA/blob/main/LOGO%20COLIBRI.png?raw=true", width=150)

st.sidebar.markdown('<h1 style="text-align: center; font-size: 2.0rem; margin-top: -10px;">Rapi Tienda<br>Acuarela</h1>', unsafe_allow_html=True)
st.sidebar.markdown("<p style='text-align: center; margin-top: -10px;'>Powered by <strong>SAVA</strong></p>", unsafe_allow_html=True)

PAGES = {
    "🏠 Inicio": "house",
    "🛰️ Escáner USB": "upc-scan",
    "📦 Inventario": "box-seam",
    "👥 Proveedores": "people",
    "🛒 Ventas": "cart4",
    "📊 Analítica": "graph-up-arrow",
    "📈 Reporte Diario": "clipboard-data",
    "🏢 Acerca de SAVA": "building"
}
for page_name, icon in PAGES.items():
    if st.sidebar.button(f"{page_name}", key=f"nav_{page_name}", width='stretch', type="primary" if st.session_state.page == page_name else "secondary"):
        st.session_state.page = page_name
        st.session_state.editing_item_id = None
        st.session_state.scanned_item_data = None
        st.session_state.usb_scan_result = None
        st.session_state.add_sku_input = ""
        st.session_state.should_clear_inventory_form = False
        st.rerun()

st.sidebar.markdown("---")
st.sidebar.markdown("<small>© 2026 SAVA & Rapi Tienda Acuarela. Todos los derechos reservados.</small>", unsafe_allow_html=True)

# --- RENDERIZADO DE PÁGINAS ---
if st.session_state.page != "🏠 Inicio":
    st.markdown(f'<h1 class="main-header">{st.session_state.page}</h1>', unsafe_allow_html=True)
    st.markdown("<hr>", unsafe_allow_html=True)

# --- PÁGINAS ---
if st.session_state.page == "🏠 Inicio":
    col_img, col_title = st.columns([1, 5])
    with col_img:
        st.image("https://github.com/GIUSEPPESAN21/LOGO-SAVA/blob/main/LOGO%20COLIBRI.png?raw=true", width=130)
    with col_title:
        st.markdown('<h1 class="main-header" style="text-align: left; margin-top: 20px;">Bienvenido a Rapi Tienda Acuarela</h1>', unsafe_allow_html=True)
        st.subheader("Sistema de gestión inteligente e inventario automatizado")

    st.markdown("""
    Este sistema transforma la manera en que gestionas tu inventario en **Rapi Tienda Acuarela**, combinando inteligencia artificial 
    de vanguardia con una interfaz intuitiva para darte control, precisión y eficiencia sin precedentes.
    """)
    st.markdown("---")

    st.subheader("Resumen del Negocio en Tiempo Real")
    items = []
    orders = []
    suppliers = []
    try:
        items = get_cached_inventory()
        orders = get_cached_orders() 
        suppliers = get_cached_suppliers()
        total_inventory_value = sum(item.get('quantity', 0) * item.get('purchase_price', 0) for item in items if isinstance(item.get('quantity'), (int, float)) and isinstance(item.get('purchase_price'), (int, float)))
        processing_orders_count = len([o for o in orders if o.get('status') == 'processing'])

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("📦 Artículos Únicos", len(items))
        c2.metric("💰 Valor del Inventario", f"${total_inventory_value:,.2f}")
        c3.metric("⏳ Ventas en Proceso", processing_orders_count)
        c4.metric("👥 Proveedores", len(suppliers))
    except Exception as e:
        st.error(f"No se pudieron cargar las estadísticas: {e}")
        items, orders, suppliers = [], [], []
    st.markdown("---")

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("Acciones Rápidas")
        if st.button("🛰️ Usar Escáner USB", width='stretch'):
             st.session_state.page = "🛰️ Escáner USB"; st.rerun()
        if st.button("📝 Crear Nueva Venta", width='stretch'):
            st.session_state.page = "🛒 Ventas"; st.rerun()
        if st.button("➕ Añadir Artículo", width='stretch'):
            st.session_state.page = "📦 Inventario"; st.rerun()

    with col2:
        st.subheader("Alertas de Stock Bajo")
        if items:
            low_stock_items = [
                item for item in items if
                item.get('min_stock_alert') is not None and isinstance(item.get('quantity'), (int, float)) and
                item['quantity'] <= item.get('min_stock_alert', 0)
            ]
            if not low_stock_items:
                st.success("¡Todo el inventario está por encima del umbral mínimo!")
            else:
                with st.container(height=200):
                    for item in low_stock_items:
                        st.warning(f"**{item.get('name', 'N/A')}**: {item.get('quantity', 0)} unidades restantes (Umbral: {item.get('min_stock_alert', 0)})")
        else:
            st.info("No hay datos de inventario para mostrar alertas.")


elif st.session_state.page == "🛰️ Escáner USB":
    st.info("Conecta tu lector de códigos de barras USB. Haz clic en el campo de texto y comienza a escanear.")

    mode = st.radio("Selecciona el modo de operación:",
                    ("Gestión de Inventario", "Punto de Venta (Salida Rápida)"),
                    horizontal=True, key="usb_scanner_mode")

    st.markdown("---")

    if mode == "Gestión de Inventario":
        col1, col2 = st.columns(2)
        with col1:
            st.subheader("Escanear para Gestionar")
            with st.form("usb_inventory_scan_form", clear_on_submit=True):
                barcode_input = st.text_input("Código de Barras", key="usb_barcode_inv_input",
                                              help="Haz clic aquí antes de escanear.")
                submitted = st.form_submit_button("Buscar / Registrar", width='stretch')
                if submitted and barcode_input:
                    st.session_state.usb_scan_result = barcode_manager.handle_inventory_scan(barcode_input)
                elif submitted and not barcode_input:
                    st.warning("Por favor, ingresa o escanea un código de barras.")

        with col2:
            st.subheader("Resultado del Escaneo")
            result = st.session_state.get('usb_scan_result')

            if not result:
                st.info("Esperando escaneo...")
            elif result['status'] == 'error':
                st.error(result['message'])
                st.session_state.usb_scan_result = None
            elif result['status'] == 'found':
                item = result['item']
                st.success(f"✔️ Producto Encontrado: **{item.get('name', 'N/A')}**")

                with st.form("update_item_form"):
                    st.write(f"**Stock Actual:** {item.get('quantity', 0)}")
                    st.write(f"**Precio de Venta:** ${item.get('sale_price', 0):.2f}")

                    new_quantity = st.number_input("Nueva Cantidad Total", min_value=0, value=item.get('quantity', 0), step=1)
                    new_price = st.number_input("Nuevo Precio de Venta ($)", min_value=0.0, value=item.get('sale_price', 0.0), format="%.2f")

                    if st.form_submit_button("Actualizar Producto", type="primary", width='stretch'):
                        updated_data = item.copy()
                        updated_data.update({
                            'quantity': int(new_quantity),
                            'sale_price': float(new_price),
                            'updated_at': datetime.now(timezone.utc).isoformat()
                        })
                        try:
                            firebase.save_inventory_item(updated_data, item['id'], is_new=False, details="Actualización vía Escáner USB.")
                            get_cached_inventory.clear() # <- Limpia caché
                            st.success(f"¡'{item.get('name', 'N/A')}' actualizado con éxito!")
                            st.session_state.usb_scan_result = None
                            st.rerun()
                        except Exception as update_e:
                             st.error(f"Error al actualizar: {update_e}")

            elif result['status'] == 'not_found':
                barcode = result['barcode']
                st.warning(f"⚠️ El código '{barcode}' no existe. Por favor, regístralo.")

                with st.form("create_from_usb_scan_form", clear_on_submit=False):
                    st.markdown(f"**Código de Barras:** `{barcode}`")
                    name = st.text_input("Nombre del Producto")
                    quantity = st.number_input("Cantidad Inicial", min_value=1, step=1, value=1)
                    sale_price = st.number_input("Precio de Venta ($)", min_value=0.0, format="%.2f", value=0.0)
                    purchase_price = st.number_input("Precio de Compra ($)", min_value=0.0, format="%.2f", value=0.0)

                    if st.form_submit_button("Guardar Nuevo Producto", type="primary", width='stretch'):
                        if name and quantity > 0:
                            data = {
                                "name": name,
                                "quantity": int(quantity),
                                "sale_price": float(sale_price),
                                "purchase_price": float(purchase_price),
                                "updated_at": datetime.now(timezone.utc).isoformat()
                             }
                            try:
                                firebase.save_inventory_item(data, barcode, is_new=True, details="Creado vía Escáner USB.")
                                get_cached_inventory.clear() # <- Limpia caché
                                st.success(f"¡Producto '{name}' guardado!")
                                st.session_state.usb_scan_result = None 
                                st.rerun()
                            except Exception as create_e:
                                st.error(f"Error al guardar: {create_e}")
                        else:
                            st.warning("El nombre y la cantidad (mayor que 0) son obligatorios.")

    elif mode == "Punto de Venta (Salida Rápida)":
        col1, col2 = st.columns([2, 3])
        with col1:
            st.subheader("Escanear Productos para Venta")
            with st.form("usb_sale_scan_form", clear_on_submit=True):
                barcode_input = st.text_input("Escanear Código de Producto", key="usb_barcode_sale_input")
                submitted = st.form_submit_button("Añadir a la Venta", width='stretch')
                if submitted and barcode_input:
                    updated_list, status_msg = barcode_manager.add_item_to_sale(barcode_input, st.session_state.usb_sale_items)
                    st.session_state.usb_sale_items = updated_list

                    if status_msg['status'] == 'success': st.toast(status_msg['message'], icon="✅")
                    elif status_msg['status'] == 'warning': st.toast(status_msg['message'], icon="⚠️")
                    else: st.error(status_msg['message'])
                    st.rerun()
                elif submitted and not barcode_input:
                     st.warning("Por favor, escanea un código de producto.")

        with col2:
            st.subheader("Detalle de la Venta Actual")
            if not st.session_state.usb_sale_items:
                st.info("Escanea un producto para comenzar...")
            else:
                total_sale_price = 0
                df_items = []
                for item in st.session_state.usb_sale_items:
                    sale_price = item.get('sale_price', 0.0)
                    quantity = item.get('quantity', 0)
                    total_item_price = sale_price * quantity
                    total_sale_price += total_item_price
                    df_items.append({
                        "Producto": item.get('name', 'N/A'),
                        "Cantidad": quantity,
                        "Precio Unit.": f"${sale_price:.2f}",
                        "Subtotal": f"${total_item_price:.2f}"
                    })

                st.dataframe(pd.DataFrame(df_items), use_container_width=True, hide_index=True)
                st.markdown(f"### Total Venta: `${total_sale_price:,.2f}`")

                # --- NUEVA SECCIÓN: FIADO Y CÁLCULO DE VUELTO ---
                st.markdown("#### Método de Pago")
                is_fiado = st.checkbox("¿Marcar como FIADO (Crédito)?", key="usb_fiado_check")
                customer_name = "Cliente General"
                cash_received_usb = 0.0
                
                if is_fiado:
                    customer_name = st.text_input("Nombre del Cliente (Deudor)", placeholder="Ej: Juan Pérez")
                    if not customer_name:
                        st.caption("⚠️ Debes ingresar un nombre para fiar.")
                else:
                    cash_received_usb = st.number_input("Efectivo Recibido ($)", min_value=0.0, step=1000.0, format="%.2f", key="usb_cash_received")
                    if cash_received_usb > 0:
                        change_usb = cash_received_usb - total_sale_price
                        if change_usb >= 0:
                            st.success(f"💰 **Cambio a devolver:** ${change_usb:,.2f}")
                        else:
                            st.error(f"⚠️ Faltan ${abs(change_usb):,.2f} para completar el pago.")

                c1, c2 = st.columns(2)
                btn_label = "✅ Finalizar Venta" if not is_fiado else "📝 Registrar Fiado"

                if c1.button(btn_label, type="primary", width='stretch'):
                    if is_fiado and not customer_name:
                         st.error("Error: Falta el nombre del cliente para fiar.")
                    elif not is_fiado and cash_received_usb > 0 and cash_received_usb < total_sale_price:
                         st.error("Error: El efectivo recibido es menor al total de la venta.")
                    else:
                        sale_id = f"VentaDirecta-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}" 
                        
                        payment_info = {
                            'method': 'fiado' if is_fiado else 'efectivo',
                            'customer': customer_name
                        }

                        try:
                            success, msg, alerts = firebase.process_direct_sale(st.session_state.usb_sale_items, sale_id, payment_info)
                            if success:
                                get_cached_inventory.clear() # <- Limpia caché
                                get_cached_orders.clear() # <- Limpia caché
                                st.success(msg)
                                send_whatsapp_alert(f"💸 Venta Rápida Procesada: {sale_id} por un total de ${total_sale_price:,.2f}")
                                for alert in alerts: send_whatsapp_alert(f"📉 ALERTA DE STOCK: {alert}")
                                st.session_state.usb_sale_items = []
                                st.rerun()
                            else:
                                st.error(msg)
                        except Exception as sale_e:
                            st.error(f"Error al procesar la venta: {sale_e}")

                if c2.button("❌ Cancelar Venta", width='stretch'):
                    st.session_state.usb_sale_items = []
                    st.toast("Venta cancelada.")
                    st.rerun()

elif st.session_state.page == "📦 Inventario":
    # Handling Editing State
    if st.session_state.editing_item_id:
        item_id_to_edit = st.session_state.editing_item_id
        try:
            item_to_edit = firebase.get_inventory_item_details(item_id_to_edit)
            if not item_to_edit:
                st.error(f"No se encontró el artículo con ID {item_id_to_edit} para editar.")
                st.session_state.editing_item_id = None
                st.rerun()
            else:
                 st.subheader(f"✏️ Editando: {item_to_edit.get('name', 'N/A')}")
                 with st.form("edit_item_form"):
                    suppliers = get_cached_suppliers()
                    supplier_map = {s.get('name', f"ID: {s.get('id')}"): s.get('id') for s in suppliers}
                    supplier_names = [""] + list(supplier_map.keys())
                    current_supplier_name = item_to_edit.get('supplier_name')
                    current_supplier_index = supplier_names.index(current_supplier_name) if current_supplier_name in supplier_names else 0

                    name = st.text_input("Nombre del Artículo", value=item_to_edit.get('name', ''))
                    quantity = st.number_input("Cantidad Actual", value=item_to_edit.get('quantity', 0), min_value=0, step=1)
                    purchase_price = st.number_input("Costo de Compra ($)", value=item_to_edit.get('purchase_price', 0.0), format="%.2f")
                    sale_price = st.number_input("Precio de Venta ($)", value=item_to_edit.get('sale_price', 0.0), format="%.2f")
                    min_stock_alert = st.number_input("Umbral de Alerta", value=item_to_edit.get('min_stock_alert', 0), min_value=0, step=1)
                    selected_supplier_name = st.selectbox("Proveedor", supplier_names, index=current_supplier_index)

                    c1, c2 = st.columns(2)
                    save_pressed = c1.form_submit_button("Guardar Cambios", type="primary", width='stretch')
                    cancel_pressed = c2.form_submit_button("Cancelar", width='stretch')

                    if save_pressed:
                        if name:
                            supplier_id = supplier_map.get(selected_supplier_name)
                            data = {
                                "name": name,
                                "quantity": int(quantity),
                                "purchase_price": float(purchase_price),
                                "sale_price": float(sale_price),
                                "min_stock_alert": int(min_stock_alert),
                                "supplier_id": supplier_id,
                                "supplier_name": selected_supplier_name if supplier_id else "",
                                "updated_at": datetime.now(timezone.utc).isoformat()
                            }
                            try:
                                firebase.save_inventory_item(data, item_id_to_edit, is_new=False, details="Edición manual de datos.")
                                get_cached_inventory.clear() # <- Limpia caché
                                st.success(f"Artículo '{name}' actualizado.")
                                st.session_state.editing_item_id = None
                                st.rerun()
                            except Exception as edit_e:
                                st.error(f"Error al guardar cambios: {edit_e}")
                        else:
                             st.warning("El nombre del artículo no puede estar vacío.")

                    if cancel_pressed:
                        st.session_state.editing_item_id = None
                        st.rerun()

        except Exception as load_e:
             st.error(f"Error al cargar datos del artículo para editar: {load_e}")
             st.session_state.editing_item_id = None 

    # Handling Display and Add Tabs
    else:
        tab1, tab2 = st.tabs(["📋 Inventario Actual", "➕ Añadir Artículo"])
        with tab1:
            search_query = st.text_input(" Buscar por Nombre o Código/ID", placeholder="Ej: Laptop, 750100100200")
            try:
                items = get_cached_inventory()

                if search_query:
                    search_query_lower = search_query.lower()
                    filtered_items = [
                        item for item in items if
                        (search_query_lower in item.get('name', '').lower()) or
                        (search_query_lower in item.get('id', '').lower())
                    ]
                else:
                    filtered_items = items

                if not filtered_items:
                    st.info("No se encontraron productos." if not search_query else "No se encontraron productos que coincidan con la búsqueda.")
                else:
                    for item in filtered_items:
                        item_id = item.get('id', 'N/A')
                        with st.container(border=True):
                            c1, c2, c3, c4, c5 = st.columns([4, 2, 2, 1, 1])
                            c1.markdown(f"**{item.get('name', 'N/A')}**")
                            c1.caption(f"ID: {item_id}")
                            c2.metric("Stock", item.get('quantity', 0))
                            c3.metric("Precio Venta", f"${item.get('sale_price', 0):,.2f}")
                            
                            if c4.button("✏️", key=f"edit_{item_id}", help="Editar este artículo"):
                                st.session_state.editing_item_id = item_id
                                st.rerun()
                            
                            if c5.button("🗑️", key=f"del_{item_id}", help="Eliminar permanentemente"):
                                show_delete_confirmation(item_id, item.get('name', 'Producto'))

            except Exception as view_e:
                 st.error(f"Error al cargar el inventario: {view_e}")

        with tab2:
            st.subheader("Añadir Nuevo Artículo al Inventario")
            
            if st.session_state.get('should_clear_inventory_form'):
                st.session_state.add_sku_input = "" 
                st.session_state.new_item_name = ""
                st.session_state.new_item_qty = 1
                st.session_state.new_item_purchase = 0.0
                st.session_state.new_item_sale = 0.0
                st.session_state.new_item_alert = 0
                st.session_state.new_item_supplier = ""
                st.session_state.should_clear_inventory_form = False

            st.markdown("##### 1️⃣ Paso 1: Escanea o Escribe el Código (SKU)")
            st.info("Usa tu lector de código de barras aquí. Si el producto no existe, podrás crearlo en el paso siguiente.")
            
            sku_candidate = st.text_input("Código del Producto", key="add_sku_input", placeholder="Escanea aquí...")

            if sku_candidate:
                try:
                    existing_item = firebase.get_inventory_item_details(sku_candidate)
                    
                    if existing_item:
                        st.warning(f"⚠️ El producto con ID **{sku_candidate}** ya existe: '{existing_item.get('name')}'.")
                        col_ex_1, col_ex_2 = st.columns(2)
                        if col_ex_1.button("✏️ Editar este producto existente", width='stretch'):
                            st.session_state.editing_item_id = sku_candidate
                            st.session_state.page = "📦 Inventario"
                            st.rerun()
                        if col_ex_2.button("🔄 Limpiar y Escanear Otro", width='stretch', on_click=set_clear_form_flag):
                            pass 
                    else:
                        st.success(f"✨ ID Disponible: **{sku_candidate}**. Completa los detalles abajo:")
                        st.markdown("---")
                        
                        st.markdown("##### 2️⃣ Paso 2: Detalles del Nuevo Producto")
                        try:
                            suppliers = get_cached_suppliers()
                            supplier_map = {s.get('name', f"ID: {s.get('id')}"): s.get('id') for s in suppliers}
                            supplier_names = [""] + list(supplier_map.keys())

                            with st.form("create_new_item_step2", clear_on_submit=False):
                                name = st.text_input("Nombre del Artículo", key="new_item_name")
                                quantity = st.number_input("Cantidad Inicial", min_value=0, step=1, key="new_item_qty")
                                purchase_price = st.number_input("Costo de Compra ($)", min_value=0.0, format="%.2f", key="new_item_purchase")
                                sale_price = st.number_input("Precio de Venta ($)", min_value=0.0, format="%.2f", key="new_item_sale")
                                min_stock_alert = st.number_input("Umbral de Alerta", min_value=0, step=1, key="new_item_alert")
                                
                                selected_supplier_name = st.selectbox("Proveedor", supplier_names, key="new_item_supplier")

                                st.form_submit_button(
                                    "💾 Guardar Producto", 
                                    type="primary", 
                                    width='stretch',
                                    on_click=save_new_item_callback,
                                    args=(supplier_map, sku_candidate)
                                )

                        except Exception as sup_e:
                            st.error(f"Error cargando proveedores: {sup_e}")

                except Exception as lookup_e:
                    st.error(f"Error verificando ID: {lookup_e}")


elif st.session_state.page == "👥 Proveedores":
    col1, col2 = st.columns([1, 2])
    with col1:
        with st.form("add_supplier_form", clear_on_submit=True):
            st.subheader("Añadir Proveedor")
            name = st.text_input("Nombre del Proveedor")
            contact = st.text_input("Persona de Contacto")
            email = st.text_input("Email")
            phone = st.text_input("Teléfono")
            if st.form_submit_button("Guardar", type="primary", width='stretch'):
                if name:
                    try:
                        firebase.add_supplier({
                            "name": name,
                            "contact_person": contact,
                            "email": email,
                            "phone": phone
                        })
                        get_cached_suppliers.clear() # <- Limpia caché
                        st.success(f"Proveedor '{name}' añadido.")
                        st.rerun() 
                    except Exception as add_sup_e:
                         st.error(f"Error al añadir proveedor: {add_sup_e}")
                else:
                    st.warning("El nombre del proveedor es obligatorio.")
    with col2:
        st.subheader("Lista de Proveedores")
        try:
            suppliers = get_cached_suppliers()
            if not suppliers:
                st.info("No hay proveedores registrados.")
            else:
                for s in suppliers:
                    with st.expander(f"**{s.get('name', 'N/A')}**"):
                        st.write(f"**Contacto:** {s.get('contact_person', 'N/A')}")
                        st.write(f"**Email:** {s.get('email', 'N/A')}")
                        st.write(f"**Teléfono:** {s.get('phone', 'N/A')}")
        except Exception as list_sup_e:
             st.error(f"Error al cargar la lista de proveedores: {list_sup_e}")


elif st.session_state.page == "🛒 Ventas":
    try:
        items_from_db = get_cached_inventory()
    except Exception as e:
        st.error(f"Error al cargar artículos de inventario: {e}")
        items_from_db = [] 

    col1, col2 = st.columns([2, 3])
    with col1:
        st.subheader("Añadir Artículos a la Venta")

        add_method = st.radio("Método para añadir:", ("Selección Manual", "Escanear para Venta"), horizontal=True, key="add_order_method")

        if add_method == "Selección Manual":
            if not items_from_db:
                st.warning("No hay artículos en el inventario para seleccionar.")
            else:
                inventory_by_name = {item['name']: item for item in items_from_db if 'name' in item}
                options = [""] + sorted(list(inventory_by_name.keys())) 
                selected_name = st.selectbox("Selecciona un artículo", options, key="manual_select_item")

                if selected_name:
                    item_to_add = inventory_by_name[selected_name]
                    item_id = item_to_add.get('id', 'N/A')
                    qty_to_add = st.number_input(f"Cantidad de '{selected_name}'", min_value=1, value=1, step=1, key=f"sel_qty_{item_id}")

                    if st.button(f"Añadir {qty_to_add} a la Venta", width='stretch'):
                        updated_order_items, status_msg = barcode_manager.add_item_to_order_list(item_to_add, st.session_state.order_items, qty_to_add)
                        st.session_state.order_items = updated_order_items
                        if status_msg['status'] == 'success':
                            st.toast(status_msg['message'], icon="✅")
                        else: 
                            st.warning(status_msg['message']) 
                        st.rerun() 

        elif add_method == "Escanear para Venta":
             with st.form("order_scan_form", clear_on_submit=True):
                barcode_input = st.text_input("Escanear Código de Producto", key="order_barcode_scan_input")
                submitted = st.form_submit_button("Buscar y Añadir", width='stretch')

                if submitted and barcode_input:
                    try:
                        item_data = firebase.get_inventory_item_details(barcode_input)
                        if item_data:
                             updated_order_items, status_msg = barcode_manager.add_item_to_order_list(item_data, st.session_state.order_items, 1)
                             st.session_state.order_items = updated_order_items
                             if status_msg['status'] == 'success':
                                 st.toast(status_msg['message'], icon="✅")
                             else:
                                 st.warning(status_msg['message'])
                        else:
                            st.error(f"El código '{barcode_input}' no fue encontrado en el inventario.")
                        st.rerun() 
                    except Exception as scan_add_e:
                        st.error(f"Error al procesar escaneo: {scan_add_e}")
                elif submitted and not barcode_input:
                    st.warning("Por favor, escanea un código.")


    with col2:
        st.subheader("Detalle de la Venta Actual")
        if not st.session_state.order_items:
            st.info("Añade artículos para comenzar una venta.")
        else:
            order_df_data = []
            for item in st.session_state.order_items:
                order_df_data.append({
                    "id": item.get('id', 'N/A'),
                    "Producto": item.get('name', 'N/A'),
                    "Cantidad": item.get('order_quantity', 1), 
                    "Precio Unit.": item.get('sale_price', 0.0),
                    "Subtotal": item.get('sale_price', 0.0) * item.get('order_quantity', 1)
                })

            order_df = pd.DataFrame(order_df_data)

            st.write("Puedes editar la cantidad directamente en la tabla:")
            edited_df = st.data_editor(
                order_df,
                column_config={
                    "id": None, 
                    "Producto": st.column_config.TextColumn(disabled=True),
                    "Cantidad": st.column_config.NumberColumn(min_value=1, step=1, required=True),
                    "Precio Unit.": st.column_config.NumberColumn(format="$%.2f", disabled=True),
                    "Subtotal": st.column_config.NumberColumn(format="$%.2f", disabled=True)
                },
                hide_index=True,
                use_container_width=True,
                key="order_editor" 
            )

            updated_items_from_editor = []
            total_price = 0
            for index, row in edited_df.iterrows():
                item_id = row['id']
                new_quantity = row['Cantidad']
                original_item = next((item for item in st.session_state.order_items if item['id'] == item_id), None)
                if original_item:
                    inventory_item = firebase.get_inventory_item_details(item_id)
                    available_stock = inventory_item.get('quantity', 0) if inventory_item else 0

                    if new_quantity > available_stock:
                         st.warning(f"Stock insuficiente para '{row['Producto']}'. Máximo disponible: {available_stock}. Ajustando a {available_stock}.")
                         new_quantity = available_stock 

                    if new_quantity > 0: 
                        original_item['order_quantity'] = new_quantity
                        subtotal = original_item.get('sale_price', 0.0) * new_quantity
                        total_price += subtotal
                        updated_items_from_editor.append(original_item)
                    else:
                        st.toast(f"'{row['Producto']}' eliminado de la venta (cantidad 0).", icon="🗑️")


            if len(updated_items_from_editor) != len(st.session_state.order_items) or any(
                st.session_state.order_items[i]['order_quantity'] != updated_items_from_editor[i]['order_quantity']
                for i in range(len(updated_items_from_editor)) if i < len(st.session_state.order_items)
            ):
                 st.session_state.order_items = updated_items_from_editor
                 st.rerun()

            st.metric("Precio Total de la Venta", f"${total_price:,.2f}")

            try:
                order_count = firebase.get_order_count()
                default_title = f"Venta #{order_count + 1}"
            except Exception as count_e:
                st.warning(f"No se pudo obtener el contador de ventas: {count_e}")
                default_title = "Nueva Venta"

            # --- NUEVA SECCIÓN: FIADO Y CÁLCULO DE VUELTO EN VENTA MANUAL ---
            st.write("---")
            is_credit = st.checkbox("Venta a Crédito (Fiado)", key="man_fiado")
            client = "Cliente General"
            cash_received_man = 0.0

            if is_credit:
                client = st.text_input("Nombre del Cliente (Deudor)", key="man_client", placeholder="Ej: Juan Pérez")
            else:
                cash_received_man = st.number_input("Efectivo Recibido ($)", min_value=0.0, step=1000.0, format="%.2f", key="man_cash_received")
                if cash_received_man > 0:
                    change_man = cash_received_man - total_price
                    if change_man >= 0:
                        st.success(f"💰 **Cambio a devolver:** ${change_man:,.2f}")
                    else:
                        st.error(f"⚠️ Faltan ${abs(change_man):,.2f} para completar el pago.")

            with st.form("order_form"):
                title = st.text_input("Nombre de la Venta (opcional)", placeholder=default_title)
                final_title = title if title else default_title
                
                if st.form_submit_button("Crear Venta", type="primary", width='stretch'):
                    if not st.session_state.order_items:
                        st.warning("No hay artículos en la venta.")
                    elif is_credit and not client:
                         st.error("Error: Falta el nombre del cliente para fiar.")
                    elif not is_credit and cash_received_man > 0 and cash_received_man < total_price:
                         st.error("Error: El efectivo recibido es menor al total de la venta.")
                    else:
                        ingredients_for_db = []
                        valid_order = True
                        for item in st.session_state.order_items:
                            inventory_item = firebase.get_inventory_item_details(item['id'])
                            available_stock = inventory_item.get('quantity', 0) if inventory_item else 0
                            if item['order_quantity'] > available_stock:
                                st.error(f"¡Stock insuficiente para '{item['name']}' al finalizar! Disponible: {available_stock}, Pedido: {item['order_quantity']}.")
                                valid_order = False
                                break 
                            ingredients_for_db.append({
                                'id': item['id'],
                                'name': item['name'],
                                'quantity': item['order_quantity']
                            })

                        if valid_order:
                            order_data = {
                                'title': final_title,
                                'price': total_price,
                                'ingredients': ingredients_for_db,
                                'status': 'processing',
                                'timestamp': datetime.now(timezone.utc),
                                'payment_method': 'fiado' if is_credit else 'efectivo',
                                'customer_name': client
                            }
                            try:
                                firebase.create_order(order_data)
                                get_cached_orders.clear() # <- Limpia caché
                                st.success(f"Venta '{final_title}' creada con éxito.")
                                send_whatsapp_alert(f"🧾 Nueva Venta: {final_title} por ${total_price:,.2f}")
                                st.session_state.order_items = []
                                st.rerun() 
                            except Exception as create_order_e:
                                st.error(f"Error al crear la venta: {create_order_e}")

    st.markdown("---")
    st.subheader("⏳ Ventas en Proceso")
    try:
        processing_orders = [o for o in get_cached_orders() if o.get('status') == 'processing']
        if not processing_orders:
            st.info("No hay ventas en proceso.")
        else:
            for order in processing_orders:
                order_id = order.get('id', 'N/A')
                with st.expander(f"**{order.get('title', 'N/A')}** - ${order.get('price', 0):,.2f}"):
                    st.write("Artículos:")
                    for item in order.get('ingredients', []):
                        st.write(f"- {item.get('name', 'N/A')} (x{item.get('quantity', 0)})")
                    c1, c2 = st.columns(2)
                    if c1.button("✅ Completar Venta", key=f"comp_{order_id}", type="primary", width='stretch'):
                        try:
                            success, msg, alerts = firebase.complete_order(order_id)
                            if success:
                                get_cached_inventory.clear() # <- Limpia caché
                                get_cached_orders.clear() # <- Limpia caché
                                st.success(msg)
                                send_whatsapp_alert(f"✅ Venta Completada: {order.get('title', 'N/A')}")
                                for alert in alerts: send_whatsapp_alert(f"📉 ALERTA DE STOCK: {alert}")
                                st.rerun() 
                            else:
                                st.error(msg)
                        except Exception as complete_e:
                             st.error(f"Error al completar venta: {complete_e}")

                    if c2.button("❌ Cancelar Venta", key=f"canc_{order_id}", width='stretch'):
                        try:
                             firebase.cancel_order(order_id)
                             get_cached_orders.clear() # <- Limpia caché
                             st.toast(f"Venta '{order.get('title', 'N/A')}' cancelada.")
                             st.rerun() 
                        except Exception as cancel_e:
                             st.error(f"Error al cancelar venta: {cancel_e}")

    except Exception as proc_ord_e:
        st.error(f"Error al cargar ventas en proceso: {proc_ord_e}")


elif st.session_state.page == "📊 Analítica":
    # LAZY LOADING: Importamos herramientas matemáticas solo si se entra a Analítica
    import plotly.express as px
    from statsmodels.tsa.holtwinters import ExponentialSmoothing

    try:
        all_orders = get_cached_orders()
        completed_orders = [o for o in all_orders if o.get('status') == 'completed']
        all_inventory_items = get_cached_inventory()
        suppliers_list = get_cached_suppliers() 
    except Exception as e:
        st.error(f"No se pudieron cargar los datos para el análisis: {e}")
        completed_orders, all_inventory_items, suppliers_list = [], [], [] 

    if not completed_orders and not all_inventory_items:
        st.info("No hay datos suficientes para generar analíticas.")
    else:
        tab1, tab2, tab3, tab4 = st.tabs(["💰 Rendimiento Financiero", "🔄 Rotación de Inventario", "📈 Predicción de Demanda", "📥 Exportar Base de Datos"])

        # Tab 1: Financial Performance
        with tab1:
            st.subheader("Indicadores Clave de Rendimiento (KPIs)")
            total_revenue = sum(o.get('price', 0) for o in completed_orders)
            total_cogs = 0
            for o in completed_orders:
                for ing in o.get('ingredients', []):
                    purchase_price = ing.get('purchase_price', 0.0)
                    quantity = ing.get('quantity', 0)
                    if isinstance(purchase_price, (int, float)) and isinstance(quantity, (int, float)):
                        total_cogs += purchase_price * quantity

            gross_profit = total_revenue - total_cogs
            num_orders = len(completed_orders)
            avg_order_value = total_revenue / num_orders if num_orders > 0 else 0
            profit_margin = (gross_profit / total_revenue) * 100 if total_revenue > 0 else 0

            kpi_cols = st.columns(5)
            kpi_cols[0].metric("Ingresos Totales", f"${total_revenue:,.2f}")
            kpi_cols[1].metric("Beneficio Bruto", f"${gross_profit:,.2f}")
            kpi_cols[2].metric("Margen de Beneficio", f"{profit_margin:.2f}%")
            kpi_cols[3].metric("Ventas Completadas", num_orders)
            kpi_cols[4].metric("Valor Promedio/Venta", f"${avg_order_value:,.2f}")
            st.markdown("---")

            st.subheader("Tendencia de Ingresos y Beneficios Diarios")
            sales_data = []
            for order in completed_orders:
                ts = order.get('timestamp_obj')
                if ts and isinstance(ts, datetime):
                     order_cogs = sum(ing.get('purchase_price', 0.0) * ing.get('quantity', 0) for ing in order.get('ingredients', []))
                     order_profit = order.get('price', 0.0) - order_cogs
                     sales_data.append({'Fecha': ts.date(), 'Ingresos': order.get('price', 0.0), 'Beneficios': order_profit})

            if sales_data:
                df_trends = pd.DataFrame(sales_data)
                df_trends['Fecha'] = pd.to_datetime(df_trends['Fecha'])
                df_daily_trends = df_trends.groupby('Fecha').agg(Ingresos=('Ingresos', 'sum'), Beneficios=('Beneficios', 'sum')).reset_index()
                fig = px.line(df_daily_trends, x='Fecha', y=['Ingresos', 'Beneficios'], title="Tendencias Diarias")
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.warning("No hay suficientes datos de fecha para generar un gráfico de tendencias.")

        # Tab 2: Inventory Rotation
        with tab2:
            all_items_sold_data = []
            for o in completed_orders:
                for ing in o.get('ingredients', []):
                    if 'name' in ing and 'id' in ing:
                        all_items_sold_data.append({
                            'id': ing['id'],
                            'name': ing['name'],
                            'quantity': ing.get('quantity', 0),
                            'sale_price': ing.get('sale_price', 0.0),
                            'purchase_price': ing.get('purchase_price', 0.0)
                        })

            if not all_items_sold_data:
                 st.info("No hay datos de artículos vendidos para analizar.")
            else:
                df_sold = pd.DataFrame(all_items_sold_data)
                df_sold['profit_per_item'] = (df_sold['sale_price'] - df_sold['purchase_price']) * df_sold['quantity']

                df_sales_summary = df_sold.groupby('name').agg(
                    Unidades_Vendidas=('quantity', 'sum'),
                    Beneficio_Generado=('profit_per_item', 'sum')
                ).reset_index()

                df_top_sales = df_sales_summary.sort_values('Unidades_Vendidas', ascending=False).head(5)
                df_top_profits = df_sales_summary.sort_values('Beneficio_Generado', ascending=False).head(5)

                col1, col2 = st.columns(2)
                with col1:
                    st.subheader("Top 5 - Artículos Más Vendidos")
                    st.dataframe(df_top_sales[['name', 'Unidades_Vendidas']].rename(columns={'name':'Artículo'}), hide_index=True, use_container_width=True)
                with col2:
                    st.subheader("Top 5 - Artículos Más Rentables")
                    st.dataframe(df_top_profits[['name', 'Beneficio_Generado']].rename(columns={'name':'Artículo'}), hide_index=True, use_container_width=True,
                                 column_config={"Beneficio_Generado": st.column_config.NumberColumn(format="$%.2f")})
                st.markdown("---")

                st.subheader("Inventario de Lenta Rotación (no vendido en los últimos 30 días)")
                thirty_days_ago = datetime.now(timezone.utc) - timedelta(days=30)
                recent_sales_ids = set(ing['id'] for o in completed_orders if o.get('timestamp_obj') and o['timestamp_obj'] > thirty_days_ago for ing in o.get('ingredients', []))

                slow_moving_items = [item for item in all_inventory_items if item.get('id') not in recent_sales_ids]

                if not slow_moving_items:
                    st.success("¡Todos los artículos han tenido movimiento en los últimos 30 días!")
                else:
                     with st.container(height=200): 
                        for item in slow_moving_items:
                            st.warning(f"- **{item.get('name', 'N/A')}** (Stock actual: {item.get('quantity', 0)})")

        # Tab 3: Demand Prediction
        with tab3:
            st.subheader("Predecir Demanda Futura de un Artículo")
            if not all_inventory_items:
                st.warning("No hay artículos en el inventario para seleccionar.")
            else:
                item_names = sorted([item.get('name', 'N/A') for item in all_inventory_items if 'name' in item])
                item_to_predict = st.selectbox("Selecciona un artículo:", [""] + item_names, key="predict_item_select")

                if item_to_predict:
                    sales_history = []
                    for order in completed_orders:
                         ts = order.get('timestamp_obj')
                         if ts and isinstance(ts, datetime):
                            for item in order.get('ingredients', []):
                                if item.get('name') == item_to_predict:
                                    sales_history.append({'date': ts, 'quantity': item.get('quantity', 0)})

                    if not sales_history:
                        st.warning("No hay historial de ventas para este artículo.")
                    else:
                        df_hist = pd.DataFrame(sales_history)
                        df_hist['date'] = pd.to_datetime(df_hist['date'])
                        df_daily_sales = df_hist.set_index('date').resample('D')['quantity'].sum().fillna(0).reset_index()

                        MIN_DAYS_FOR_SEASONAL = 14 
                        MIN_DAYS_FOR_SIMPLE = 5

                        if len(df_daily_sales) < MIN_DAYS_FOR_SIMPLE:
                            st.warning(f"No hay suficientes datos ({len(df_daily_sales)} días). Se necesitan al menos {MIN_DAYS_FOR_SIMPLE} días de ventas para una predicción básica.")
                        else:
                            try:
                                sales_ts = df_daily_sales.set_index('date')['quantity']
                                model = None
                                model_info = ""

                                if len(sales_ts) >= MIN_DAYS_FOR_SEASONAL:
                                    try:
                                        model = ExponentialSmoothing(sales_ts, seasonal='add', seasonal_periods=7, trend='add', initialization_method='estimated').fit()
                                        model_info = "Modelo: Suavizado Exponencial con Tendencia y Estacionalidad (7 días)."
                                    except ValueError as seasonal_error:
                                         st.warning(f"No se pudo ajustar modelo estacional ({seasonal_error}). Intentando modelo simple.")
                                         model = None 

                                if model is None:
                                    model = ExponentialSmoothing(sales_ts, trend='add', initialization_method='estimated').fit()
                                    model_info = "Modelo: Suavizado Exponencial con Tendencia (Simple)."

                                st.info(model_info)
                                forecast_periods = 30
                                prediction = model.forecast(forecast_periods)
                                prediction[prediction < 0] = 0 

                                total_predicted_demand = int(round(prediction.sum()))
                                st.success(f"Se estima una demanda de **{total_predicted_demand} unidades** para los próximos {forecast_periods} días.")

                                forecast_dates = pd.date_range(start=sales_ts.index.max() + timedelta(days=1), periods=forecast_periods)
                                df_forecast = pd.DataFrame({'Fecha': forecast_dates, 'Predicción': prediction})

                                fig_pred = px.line(df_daily_sales, x='date', y='quantity', title=f'Historial de Ventas y Predicción para {item_to_predict}', labels={'date':'Fecha', 'quantity':'Ventas Históricas'})
                                fig_pred.add_scatter(x=df_forecast['Fecha'], y=df_forecast['Predicción'], mode='lines', name='Predicción', line=dict(dash='dash'))
                                st.plotly_chart(fig_pred, use_container_width=True)

                            except Exception as e:
                                st.error(f"No se pudo generar la predicción: {e}")

        # Tab 4: Exportación de Datos
        with tab4:
            st.subheader("Descarga Masiva de Base de Datos")
            st.markdown("""
            Esta herramienta genera un Excel con la estructura exacta de tus colecciones de base de datos:
            - **inventory**: Tu inventario completo.
            - **orders**: Todas las ventas (completadas, en proceso, etc.).
            - **orders_items**: Detalle de productos por cada orden.
            - **suppliers**: Lista de proveedores.
            """)
            
            col_down_1, col_down_2 = st.columns([1, 2])
            with col_down_1:
                if st.button("📥 Generar Excel Maestro", type="primary", use_container_width=True):
                    with st.spinner("Compilando datos y generando archivo Excel..."):
                        try:
                            df_inventory = pd.DataFrame(all_inventory_items)
                            df_suppliers = pd.DataFrame(suppliers_list)
                            all_orders_raw = get_cached_orders()
                            
                            sales_summary_data = []
                            sales_detailed_data = []

                            for o in all_orders_raw:
                                ts = o.get('timestamp_obj') 
                                if not ts: 
                                    ts = o.get('timestamp')
                                ts_str = ts.strftime('%Y-%m-%d %H:%M:%S') if ts and isinstance(ts, (datetime, pd.Timestamp)) else str(ts)
                                
                                sales_summary_data.append({
                                    'id': o.get('id'),
                                    'timestamp': ts_str,
                                    'title': o.get('title'),
                                    'price': o.get('price'),
                                    'payment_method': o.get('payment_method', 'efectivo'),
                                    'customer_name': o.get('customer_name', ''),
                                    'status': o.get('status'),
                                    'completed_at': str(o.get('completed_at', ''))
                                })

                                for item in o.get('ingredients', []):
                                    sales_detailed_data.append({
                                        'order_id': o.get('id'),
                                        'order_date': ts_str,
                                        'item_name': item.get('name'),
                                        'quantity': item.get('quantity'),
                                        'sale_price': item.get('sale_price'),
                                        'purchase_price': item.get('purchase_price', 0),
                                        'subtotal': item.get('sale_price', 0) * item.get('quantity', 0)
                                    })

                            df_orders = pd.DataFrame(sales_summary_data)
                            df_orders_items = pd.DataFrame(sales_detailed_data)

                            output = io.BytesIO()
                            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                if not df_inventory.empty:
                                    df_inventory.to_excel(writer, sheet_name='inventory', index=False)
                                if not df_orders.empty:
                                    df_orders.to_excel(writer, sheet_name='orders', index=False)
                                if not df_orders_items.empty:
                                    df_orders_items.to_excel(writer, sheet_name='orders_items', index=False)
                                if not df_suppliers.empty:
                                    df_suppliers.to_excel(writer, sheet_name='suppliers', index=False)
                            
                            output.seek(0)
                            st.session_state['excel_buffer'] = output
                            st.toast("Archivo Excel generado con estructura de base de datos.", icon="✅")

                        except Exception as e:
                            st.error(f"Error al generar el Excel: {e}")

            if 'excel_buffer' in st.session_state:
                file_name = f"SAVA_DB_Export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                st.download_button(
                    label="⬇️ Click aquí para descargar el archivo",
                    data=st.session_state['excel_buffer'],
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )


elif st.session_state.page == "📈 Reporte Diario":
    # ── Zona horaria Colombia (UTC-5) ──────────────────────────────────────
    _COL_TZ = timezone(timedelta(hours=-5))
    _now_col = datetime.now(_COL_TZ)
    _MESES_ES = {
        1: "enero", 2: "febrero", 3: "marzo", 4: "abril",
        5: "mayo", 6: "junio", 7: "julio", 8: "agosto",
        9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre"
    }
    _fecha_legible = f"{_now_col.day} de {_MESES_ES[_now_col.month]} de {_now_col.year}"
    _hora_legible  = _now_col.strftime("%I:%M %p").lstrip("0")
    _mes_actual    = f"{_MESES_ES[_now_col.month].capitalize()} {_now_col.year}"

    st.subheader("📊 Reporte Diario de Ventas")
    st.markdown(f"📅 **Fecha:** {_fecha_legible} — 🕐 **Hora:** {_hora_legible}")
    st.markdown("---")

    # ── Helper: genera Excel profesional con estilos ───────────────────────
    def _build_professional_excel(
        fecha_legible, hora_legible, mes_actual,
        # Datos del día
        completed_today, items_sold_dict, df_items_today,
        total_rev, total_cash, total_credit, n_trans, avg_tick,
        total_cogs_day, gross_profit_day,
        # Datos del mes
        completed_month, month_revenue, month_cogs, month_profit,
        month_margin, month_num_sales, month_items_dict,
    ):
        '''Crea y devuelve un BytesIO con el Excel completo y estilizado.'''
        from openpyxl import Workbook
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side, GradientFill
        )
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # ── Paleta de colores corporativa ──────────────────────────────────
        COLOR_VERDE_OSC  = "1B5E20"   # verde oscuro encabezados
        COLOR_VERDE_MED  = "2E7D32"   # verde medio subencabezados
        COLOR_VERDE_CLAR = "C8E6C9"   # verde claro filas alternadas
        COLOR_NARANJA    = "E65100"   # naranja alertas
        COLOR_AZUL       = "1565C0"   # azul métricas
        COLOR_GRIS       = "ECEFF1"   # gris fondo secundario
        COLOR_BLANCO     = "FFFFFF"
        COLOR_NEGRO      = "212121"
        COLOR_AMARILLO   = "F9A825"   # fiado/crédito

        def _fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def _font(bold=False, color=COLOR_NEGRO, size=11, italic=False):
            return Font(bold=bold, color=color, size=size, italic=italic,
                        name="Calibri")

        def _border():
            thin = Side(style="thin", color="B0BEC5")
            return Border(left=thin, right=thin, top=thin, bottom=thin)

        def _align(h="center", v="center", wrap=False):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

        def _header_row(ws, row, values, fill_color, font_color="FFFFFF",
                        font_size=11, bold=True, height=22):
            '''Escribe y estiliza una fila de encabezado.'''
            ws.row_dimensions[row].height = height
            for col, val in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.fill      = _fill(fill_color)
                c.font      = _font(bold=bold, color=font_color, size=font_size)
                c.alignment = _align("center")
                c.border    = _border()

        def _data_row(ws, row, values, alt=False, fmt_map=None, height=18):
            '''Escribe y estiliza una fila de datos (color alternado).'''
            ws.row_dimensions[row].height = height
            bg = COLOR_VERDE_CLAR if alt else COLOR_BLANCO
            for col, val in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.fill      = _fill(bg)
                c.font      = _font(size=10)
                c.alignment = _align("center")
                c.border    = _border()
                if fmt_map and col in fmt_map:
                    c.number_format = fmt_map[col]

        def _auto_col_width(ws, extra=4):
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + extra, 45)

        def _merge_title(ws, row, text, end_col, fill_color, font_size=12,
                         font_color="FFFFFF", height=26):
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=end_col)
            c = ws.cell(row=row, column=1, value=text)
            c.fill      = _fill(fill_color)
            c.font      = _font(bold=True, color=font_color, size=font_size)
            c.alignment = _align("center", wrap=True)
            c.border    = _border()
            ws.row_dimensions[row].height = height

        # ══════════════════════════════════════════════════════════════════
        #  HOJA 1 — PORTADA
        # ══════════════════════════════════════════════════════════════════
        ws0 = wb.active
        ws0.title = "📋 Portada"
        ws0.sheet_view.showGridLines = False

        # Título principal
        ws0.merge_cells("A1:F1")
        c = ws0["A1"]
        c.value     = "RAPI TIENDA ACUARELA"
        c.fill      = _fill(COLOR_VERDE_OSC)
        c.font      = _font(bold=True, color=COLOR_BLANCO, size=20)
        c.alignment = _align("center")
        ws0.row_dimensions[1].height = 40

        ws0.merge_cells("A2:F2")
        c = ws0["A2"]
        c.value     = "Sistema SAVA — Reporte Completo de Gestión"
        c.fill      = _fill(COLOR_VERDE_MED)
        c.font      = _font(bold=True, color=COLOR_BLANCO, size=13)
        c.alignment = _align("center")
        ws0.row_dimensions[2].height = 26

        ws0.row_dimensions[3].height = 10

        cover_data = [
            ("📅 Fecha del Reporte", fecha_legible),
            ("🕐 Hora de Generación", hora_legible),
            ("📆 Mes en Análisis", mes_actual),
            ("🧾 Ventas Completadas Hoy", n_trans),
            ("💰 Ingresos del Día", f"${total_rev:,.2f}"),
            ("📦 Ingresos del Mes", f"${month_revenue:,.2f}"),
            ("📈 Beneficio Bruto del Mes", f"${month_profit:,.2f}"),
            ("📊 Margen del Mes", f"{month_margin:.1f}%"),
            ("🏪 Generado por", "SAVA Software for Engineering"),
        ]
        for i, (label, val) in enumerate(cover_data, 4):
            ws0.row_dimensions[i].height = 22
            cL = ws0.cell(row=i, column=1, value=label)
            cL.fill      = _fill(COLOR_GRIS)
            cL.font      = _font(bold=True, size=11)
            cL.alignment = _align("left")
            cL.border    = _border()

            cV = ws0.cell(row=i, column=2, value=val)
            cV.fill      = _fill(COLOR_BLANCO)
            cV.font      = _font(size=11)
            cV.alignment = _align("left")
            cV.border    = _border()

        ws0.column_dimensions["A"].width = 32
        ws0.column_dimensions["B"].width = 30

        # ══════════════════════════════════════════════════════════════════
        #  HOJA 2 — RESUMEN DEL DÍA (KPIs)
        # ══════════════════════════════════════════════════════════════════
        ws1 = wb.create_sheet("📊 Resumen Diario")
        ws1.sheet_view.showGridLines = False

        _merge_title(ws1, 1, f"RESUMEN DEL DÍA — {fecha_legible}", 4,
                     COLOR_VERDE_OSC, 14)

        _header_row(ws1, 2, ["Indicador", "Valor", "Detalle", "Estado"],
                    COLOR_VERDE_MED)

        kpi_rows = [
            ("💰 Ingresos Totales",       total_rev,
             "Suma de todas las ventas completadas hoy",
             "✅ Activo"),
            ("💵 Ingresos en Efectivo",   total_cash,
             "Ventas pagadas en efectivo",
             "✅ Cobrado"),
            ("📝 Ventas a Crédito/Fiado", total_credit,
             "Ventas pendientes de cobro",
             "⚠️ Por cobrar" if total_credit > 0 else "✅ Sin fiado"),
            ("📦 Costo de Mercancía",     total_cogs_day,
             "Costo de compra de productos vendidos",
             "📊 Calculado"),
            ("📈 Beneficio Bruto",        gross_profit_day,
             "Ingresos - Costo de mercancía",
             "✅ Positivo" if gross_profit_day >= 0 else "⚠️ Negativo"),
            ("🧾 N° de Transacciones",    n_trans,
             "Total de ventas completadas hoy",
             "📊 Calculado"),
            ("🎯 Ticket Promedio",        avg_tick,
             "Ingreso promedio por venta",
             "📊 Calculado"),
        ]
        margin_day = (gross_profit_day / total_rev * 100) if total_rev > 0 else 0
        kpi_rows.append(
            ("📊 Margen de Beneficio Día", f"{margin_day:.1f}%",
             "Beneficio sobre ingresos del día",
             "✅ Saludable" if margin_day >= 20 else "⚠️ Bajo"),
        )

        for i, (ind, val, det, est) in enumerate(kpi_rows, 3):
            alt = (i % 2 == 0)
            bg = COLOR_VERDE_CLAR if alt else COLOR_BLANCO
            ws1.row_dimensions[i].height = 20

            for col, v in enumerate([ind, val, det, est], 1):
                c = ws1.cell(row=i, column=col, value=v)
                c.fill      = _fill(bg)
                c.font      = _font(size=10,
                                    bold=(col == 1),
                                    color="1B5E20" if col == 1 else COLOR_NEGRO)
                c.alignment = _align("left" if col in (1, 3) else "center")
                c.border    = _border()
                if col == 2 and isinstance(v, (int, float)) and not isinstance(v, str):
                    c.number_format = '$#,##0.00'

        _auto_col_width(ws1, extra=5)

        # ══════════════════════════════════════════════════════════════════
        #  HOJA 3 — VENTAS DETALLADAS DEL DÍA
        # ══════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet("🧾 Ventas Detalladas")
        ws2.sheet_view.showGridLines = False

        _merge_title(ws2, 1, f"VENTAS DETALLADAS — {fecha_legible}", 8,
                     COLOR_VERDE_OSC, 14)

        headers_v = ["#", "Nombre de Venta", "Hora", "Cliente",
                     "Método de Pago", "N° Productos", "Total ($)", "Estado"]
        _header_row(ws2, 2, headers_v, COLOR_VERDE_MED)

        fmt_v = {7: '$#,##0.00'}
        for i, order in enumerate(completed_today, 1):
            ca = order.get('completed_at')
            hora_v = (ca.astimezone(_COL_TZ).strftime("%I:%M %p").lstrip("0")
                      if ca and isinstance(ca, datetime) else "N/A")
            metodo = ("💳 Fiado/Crédito"
                      if order.get('payment_method') == 'fiado'
                      else "💵 Efectivo")
            n_prod  = sum(it.get('quantity', 0)
                         for it in order.get('ingredients', []))
            vals = [
                i,
                order.get('title', 'Sin título'),
                hora_v,
                order.get('customer_name', 'Cliente General'),
                metodo,
                n_prod,
                order.get('price', 0),
                "✅ Completada",
            ]
            _data_row(ws2, i + 2, vals, alt=(i % 2 == 0), fmt_map=fmt_v)

        _auto_col_width(ws2, extra=4)

        # ══════════════════════════════════════════════════════════════════
        #  HOJA 4 — PRODUCTOS VENDIDOS HOY
        # ══════════════════════════════════════════════════════════════════
        ws3 = wb.create_sheet("📦 Productos Hoy")
        ws3.sheet_view.showGridLines = False

        _merge_title(ws3, 1, f"PRODUCTOS VENDIDOS HOY — {fecha_legible}", 7,
                     COLOR_VERDE_OSC, 14)

        headers_p = ["Producto", "Unidades\nVendidas", "Precio\nUnitario ($)",
                     "Costo\nUnitario ($)", "Subtotal\nVentas ($)",
                     "Costo\nTotal ($)", "Margen\nBruto ($)"]
        _header_row(ws3, 2, headers_p, COLOR_VERDE_MED, height=30)

        fmt_p = {3: '$#,##0.00', 4: '$#,##0.00', 5: '$#,##0.00',
                 6: '$#,##0.00', 7: '$#,##0.00'}

        prod_rows = []
        for _order in completed_today:
            for _ing in _order.get('ingredients', []):
                pname = _ing.get('name', 'N/A')
                qty   = _ing.get('quantity', 0)
                sp    = _ing.get('sale_price', 0.0)
                pp    = _ing.get('purchase_price', 0.0)
                if pname not in {r[0] for r in prod_rows}:
                    prod_rows.append([pname, qty, sp, pp,
                                      qty * sp, qty * pp,
                                      qty * (sp - pp)])
                else:
                    for r in prod_rows:
                        if r[0] == pname:
                            r[1] += qty
                            r[4] += qty * sp
                            r[5] += qty * pp
                            r[6] += qty * (sp - pp)
                            break

        prod_rows.sort(key=lambda r: r[4], reverse=True)

        i = 2
        for i_idx, row in enumerate(prod_rows, 3):
            i = i_idx
            _data_row(ws3, i, row, alt=(i % 2 == 0), fmt_map=fmt_p)

        # Fila de TOTALES
        tot_row = i + 1 if prod_rows else 3
        ws3.row_dimensions[tot_row].height = 22
        tot_vals = [
            "TOTALES",
            sum(r[1] for r in prod_rows),
            "",
            "",
            sum(r[4] for r in prod_rows),
            sum(r[5] for r in prod_rows),
            sum(r[6] for r in prod_rows),
        ]
        for col, v in enumerate(tot_vals, 1):
            c = ws3.cell(row=tot_row, column=col, value=v)
            c.fill      = _fill(COLOR_VERDE_OSC)
            c.font      = _font(bold=True, color=COLOR_BLANCO, size=11)
            c.alignment = _align("center")
            c.border    = _border()
            if col in fmt_p and isinstance(v, (int, float)) and not isinstance(v, str):
                c.number_format = '$#,##0.00'

        _auto_col_width(ws3, extra=4)

        # ══════════════════════════════════════════════════════════════════
        #  HOJA 5 — CLIENTES CON FIADO
        # ══════════════════════════════════════════════════════════════════
        ws4 = wb.create_sheet("📝 Clientes Fiado")
        ws4.sheet_view.showGridLines = False

        _merge_title(ws4, 1, f"CLIENTES CON CRÉDITO/FIADO — {fecha_legible}",
                     5, COLOR_NARANJA, 13)

        headers_f = ["Cliente", "Venta", "Hora", "Total ($)", "Estado"]
        _header_row(ws4, 2, headers_f, COLOR_NARANJA)

        credit_orders = [o for o in completed_today
                         if o.get('payment_method') == 'fiado']
        fmt_f = {4: '$#,##0.00'}

        if credit_orders:
            for i, order in enumerate(credit_orders, 3):
                ca = order.get('completed_at')
                hora_f = (ca.astimezone(_COL_TZ).strftime("%I:%M %p").lstrip("0")
                          if ca and isinstance(ca, datetime) else "N/A")
                vals_f = [
                    order.get('customer_name', 'Sin nombre'),
                    order.get('title', 'Sin título'),
                    hora_f,
                    order.get('price', 0),
                    "⚠️ Pendiente de cobro",
                ]
                _data_row(ws4, i, vals_f, alt=(i % 2 == 0), fmt_map=fmt_f)

            # Total fiado
            tf_row = len(credit_orders) + 3
            ws4.row_dimensions[tf_row].height = 22
            for col in range(1, 6):
                c = ws4.cell(row=tf_row, column=col)
                c.fill   = _fill(COLOR_NARANJA)
                c.font   = _font(bold=True, color=COLOR_BLANCO, size=11)
                c.border = _border()
                c.alignment = _align("center")
            ws4.cell(row=tf_row, column=1).value = "TOTAL FIADO"
            tc = ws4.cell(row=tf_row, column=4,
                          value=sum(o.get('price', 0) for o in credit_orders))
            tc.number_format = '$#,##0.00'
        else:
            ws4.merge_cells("A3:E3")
            c = ws4["A3"]
            c.value     = "✅ No hay ventas a crédito/fiado en esta fecha"
            c.fill      = _fill(COLOR_VERDE_CLAR)
            c.font      = _font(bold=True, color="1B5E20", size=11)
            c.alignment = _align("center")

        _auto_col_width(ws4, extra=4)

        # ══════════════════════════════════════════════════════════════════
        #  HOJA 6 — RESUMEN MENSUAL (KPIs)
        # ══════════════════════════════════════════════════════════════════
        ws5 = wb.create_sheet("📅 Resumen Mensual")
        ws5.sheet_view.showGridLines = False

        _merge_title(ws5, 1, f"RESUMEN MENSUAL — {mes_actual}", 4,
                     COLOR_AZUL, 14, font_color=COLOR_BLANCO)

        _header_row(ws5, 2, ["Indicador", "Valor", "Detalle", "Estado"],
                    "1565C0")

        month_avg = month_revenue / max(_now_col.day, 1)
        day_pct   = (total_rev / month_revenue * 100) if month_revenue > 0 else 0

        kpi_month = [
            ("💰 Ingresos Totales del Mes",  month_revenue,
             f"Acumulado al {fecha_legible}", "✅"),
            ("📦 Costo Total de Mercancía",  month_cogs,
             "Costo de compra acumulado", "📊"),
            ("📈 Beneficio Bruto del Mes",   month_profit,
             "Ingresos - Costos",
             "✅ Positivo" if month_profit >= 0 else "⚠️ Negativo"),
            ("📊 Margen de Beneficio",
             f"{month_margin:.2f}%",
             "Rentabilidad del mes",
             "✅ Bueno" if month_margin >= 20 else "⚠️ Bajo"),
            ("🧾 Total de Ventas del Mes",   month_num_sales,
             "Ventas completadas en el mes", "📊"),
            ("📆 Promedio Diario de Ventas", month_avg,
             "Ingreso promedio por día trabajado", "📊"),
            ("⚡ Participación del Día Hoy",
             f"{day_pct:.1f}%",
             f"Hoy aportó {day_pct:.1f}% del ingreso mensual", "📊"),
        ]
        for i, (ind, val, det, est) in enumerate(kpi_month, 3):
            alt = (i % 2 == 0)
            bg  = "BBDEFB" if alt else COLOR_BLANCO
            ws5.row_dimensions[i].height = 20
            for col, v in enumerate([ind, val, det, est], 1):
                c = ws5.cell(row=i, column=col, value=v)
                c.fill      = _fill(bg)
                c.font      = _font(size=10, bold=(col == 1),
                                    color="1565C0" if col == 1 else COLOR_NEGRO)
                c.alignment = _align("left" if col in (1, 3) else "center")
                c.border    = _border()
                if col == 2 and isinstance(v, (int, float)) and not isinstance(v, str):
                    c.number_format = '$#,##0.00'

        _auto_col_width(ws5, extra=5)

        # ══════════════════════════════════════════════════════════════════
        #  HOJA 7 — PRODUCTOS VENDIDOS DEL MES
        # ══════════════════════════════════════════════════════════════════
        ws6 = wb.create_sheet("🏆 Productos del Mes")
        ws6.sheet_view.showGridLines = False

        _merge_title(ws6, 1, f"ANÁLISIS DE PRODUCTOS — {mes_actual}", 8,
                     COLOR_AZUL, 14, font_color=COLOR_BLANCO)

        headers_m = ["Ranking", "Producto",
                     "Unidades\nVendidas", "Precio\nUnitario ($)",
                     "Ingresos\nTotal ($)", "Costo\nTotal ($)",
                     "Margen\nBruto ($)", "Margen\n(%)"]
        _header_row(ws6, 2, headers_m, "1565C0", height=30)

        # Recalcular con datos de costo del mes
        month_prod = {}
        for _order in completed_month:
            for _ing in _order.get('ingredients', []):
                pname = _ing.get('name', 'N/A')
                qty   = _ing.get('quantity', 0)
                sp    = _ing.get('sale_price', 0.0)
                pp    = _ing.get('purchase_price', 0.0)
                if pname in month_prod:
                    month_prod[pname]['qty']      += qty
                    month_prod[pname]['ingresos'] += qty * sp
                    month_prod[pname]['costo']    += qty * pp
                    month_prod[pname]['margen']   += qty * (sp - pp)
                else:
                    month_prod[pname] = {
                        'producto': pname, 'qty': qty,
                        'precio': sp,
                        'ingresos': qty * sp, 'costo': qty * pp,
                        'margen': qty * (sp - pp)
                    }

        sorted_month = sorted(month_prod.values(),
                              key=lambda x: x['ingresos'], reverse=True)

        fmt_m = {4: '$#,##0.00', 5: '$#,##0.00', 6: '$#,##0.00',
                 7: '$#,##0.00', 8: '0.0"%"'}
        for i, prod in enumerate(sorted_month, 3):
            mg_pct = (prod['margen'] / prod['ingresos'] * 100
                      if prod['ingresos'] > 0 else 0)
            vals_m = [
                i - 2,
                prod['producto'],
                prod['qty'],
                prod['precio'],
                prod['ingresos'],
                prod['costo'],
                prod['margen'],
                round(mg_pct, 1),
            ]
            _data_row(ws6, i, vals_m, alt=(i % 2 == 0), fmt_map=fmt_m)

        # Totales mes
        if sorted_month:
            tot_m = len(sorted_month) + 3
            ws6.row_dimensions[tot_m].height = 22
            tot_m_vals = [
                "", "TOTALES DEL MES",
                sum(p['qty']      for p in sorted_month),
                "",
                sum(p['ingresos'] for p in sorted_month),
                sum(p['costo']    for p in sorted_month),
                sum(p['margen']   for p in sorted_month),
                "",
            ]
            for col, v in enumerate(tot_m_vals, 1):
                c = ws6.cell(row=tot_m, column=col, value=v)
                c.fill      = _fill(COLOR_AZUL)
                c.font      = _font(bold=True, color=COLOR_BLANCO, size=11)
                c.alignment = _align("center")
                c.border    = _border()
                if col in fmt_m and isinstance(v, (int, float)) and not isinstance(v, str):
                    c.number_format = '$#,##0.00'

        _auto_col_width(ws6, extra=4)

        # ── Guardar en buffer ──────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ── Fin helper ─────────────────────────────────────────────────────────

    try:
        # ── Rangos de fecha ────────────────────────────────────────────────
        today_utc      = datetime.now(timezone.utc).date()
        start_of_day   = datetime(today_utc.year, today_utc.month,
                                  today_utc.day, tzinfo=timezone.utc)
        end_of_day     = start_of_day + timedelta(days=1)
        start_of_month = datetime(today_utc.year, today_utc.month, 1,
                                  tzinfo=timezone.utc)

        all_orders = get_cached_orders()

        completed_orders_today = [
            o for o in all_orders
            if o.get('status') == 'completed'
            and o.get('completed_at')
            and start_of_day <= o.get('completed_at') < end_of_day
        ]
        completed_orders_month = [
            o for o in all_orders
            if o.get('status') == 'completed'
            and o.get('completed_at')
            and start_of_month <= o.get('completed_at') < end_of_day
        ]

        # ── Pre-calcular datos del día ─────────────────────────────────────
        total_revenue   = sum(o.get('price', 0) for o in completed_orders_today)
        total_cash      = sum(o.get('price', 0) for o in completed_orders_today
                              if o.get('payment_method') != 'fiado')
        total_credit    = sum(o.get('price', 0) for o in completed_orders_today
                              if o.get('payment_method') == 'fiado')
        num_transactions = len(completed_orders_today)
        avg_ticket      = total_revenue / num_transactions if num_transactions > 0 else 0

        # Costo y margen del día
        total_cogs_day = 0.0
        for _o in completed_orders_today:
            for _ing in _o.get('ingredients', []):
                _pp = _ing.get('purchase_price', 0.0)
                _qq = _ing.get('quantity', 0)
                if isinstance(_pp, (int, float)) and isinstance(_qq, (int, float)):
                    total_cogs_day += _pp * _qq
        gross_profit_day = total_revenue - total_cogs_day
        margin_day_pct   = (gross_profit_day / total_revenue * 100) if total_revenue > 0 else 0

        # Consolidar productos vendidos hoy
        items_sold = {}
        for _order in completed_orders_today:
            for _item in _order.get('ingredients', []):
                _name = _item.get('name', 'N/A')
                _qty  = _item.get('quantity', 0)
                _sp   = _item.get('sale_price', 0.0)
                _pp   = _item.get('purchase_price', 0.0)
                if _name in items_sold:
                    items_sold[_name]['Cantidad']  += _qty
                    items_sold[_name]['Subtotal']  += _qty * _sp
                    items_sold[_name]['Costo Tot.'] += _qty * _pp
                    items_sold[_name]['Margen ($)'] += _qty * (_sp - _pp)
                else:
                    items_sold[_name] = {
                        'Producto':   _name,
                        'Cantidad':   _qty,
                        'P. Venta':   _sp,
                        'P. Compra':  _pp,
                        'Subtotal':   _qty * _sp,
                        'Costo Tot.': _qty * _pp,
                        'Margen ($)': _qty * (_sp - _pp),
                    }

        df_items_day = None
        if items_sold:
            df_items_day = pd.DataFrame(list(items_sold.values()))
            df_items_day = df_items_day.sort_values('Subtotal', ascending=False)

        # ── Pre-calcular datos del mes ─────────────────────────────────────
        month_revenue   = sum(o.get('price', 0) for o in completed_orders_month)
        month_cogs      = 0.0
        for _o in completed_orders_month:
            for _ing in _o.get('ingredients', []):
                _pp = _ing.get('purchase_price', 0.0)
                _qq = _ing.get('quantity', 0)
                if isinstance(_pp, (int, float)) and isinstance(_qq, (int, float)):
                    month_cogs += _pp * _qq
        month_profit    = month_revenue - month_cogs
        month_margin    = (month_profit / month_revenue * 100) if month_revenue > 0 else 0
        month_num_sales = len(completed_orders_month)

        month_items_dict = {}
        for _o in completed_orders_month:
            for _ing in _o.get('ingredients', []):
                _n = _ing.get('name', 'N/A')
                _q = _ing.get('quantity', 0)
                _s = _ing.get('sale_price', 0.0)
                if _n in month_items_dict:
                    month_items_dict[_n]['Unidades'] += _q
                    month_items_dict[_n]['Ingresos'] += _q * _s
                else:
                    month_items_dict[_n] = {
                        'Producto': _n, 'Unidades': _q, 'Ingresos': _q * _s
                    }

        # ══════════════════════════════════════════════════════════════════
        #  TABS
        # ══════════════════════════════════════════════════════════════════
        tab_dia, tab_mes = st.tabs(["📊 Reporte del Día", "📅 Análisis Mensual"])

        # ────────────────────────────────────────────────────────────────
        #  TAB 1 — REPORTE DEL DÍA
        # ────────────────────────────────────────────────────────────────
        with tab_dia:
            if not completed_orders_today:
                st.warning(f"No hay ventas registradas para hoy ({today_utc.strftime('%d/%m/%Y')}).")
            else:
                # ── 1. KPIs (6 columnas) ──────────────────────────────
                k1, k2, k3, k4, k5, k6 = st.columns(6)
                k1.metric("💰 Total Vendido",     f"${total_revenue:,.2f}")
                k2.metric("💵 Efectivo",           f"${total_cash:,.2f}")
                k3.metric("📝 Fiado",              f"${total_credit:,.2f}")
                k4.metric("📦 Costo Mercancía",    f"${total_cogs_day:,.2f}")
                k5.metric("📈 Beneficio Bruto",    f"${gross_profit_day:,.2f}")
                k6.metric("🎯 Margen del Día",     f"{margin_day_pct:.1f}%")

                st.markdown("---")
                c_l, c_r = st.columns(2)
                c_l.metric("🧾 Transacciones", num_transactions)
                c_r.metric("🎯 Ticket Promedio", f"${avg_ticket:,.2f}")

                st.markdown("---")

                # ── 2. Productos vendidos hoy ─────────────────────────
                st.subheader("📦 Productos Vendidos Hoy")
                if df_items_day is not None:
                    st.dataframe(
                        df_items_day,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "P. Venta":   st.column_config.NumberColumn(format="$%.2f"),
                            "P. Compra":  st.column_config.NumberColumn(format="$%.2f"),
                            "Subtotal":   st.column_config.NumberColumn(format="$%.2f"),
                            "Costo Tot.": st.column_config.NumberColumn(format="$%.2f"),
                            "Margen ($)": st.column_config.NumberColumn(format="$%.2f"),
                        }
                    )
                else:
                    st.info("No hay productos registrados en las ventas de hoy.")

                st.markdown("---")

                # ── 3. Desglose por método de pago ────────────────────
                st.subheader("💳 Desglose por Método de Pago")
                cash_orders   = [o for o in completed_orders_today
                                 if o.get('payment_method') != 'fiado']
                credit_orders = [o for o in completed_orders_today
                                 if o.get('payment_method') == 'fiado']
                df_pay = pd.DataFrame([
                    {"Método": "💵 Efectivo",
                     "N° Ventas": len(cash_orders),
                     "Total ($)": sum(o.get('price', 0) for o in cash_orders)},
                    {"Método": "📝 Fiado/Crédito",
                     "N° Ventas": len(credit_orders),
                     "Total ($)": sum(o.get('price', 0) for o in credit_orders)},
                ])
                st.dataframe(df_pay, use_container_width=True, hide_index=True,
                             column_config={"Total ($)": st.column_config.NumberColumn(format="$%.2f")})

                if credit_orders:
                    st.caption("📋 **Clientes con ventas fiadas hoy:**")
                    for co in credit_orders:
                        st.markdown(
                            f"- **{co.get('customer_name','Sin nombre')}**: "
                            f"${co.get('price',0):,.2f}")

                st.markdown("---")

                # ── 4. Detalle de transacciones ───────────────────────
                st.subheader("📝 Detalle de Transacciones del Día")
                for idx, order in enumerate(completed_orders_today, 1):
                    ca = order.get('completed_at')
                    order_time = (ca.astimezone(_COL_TZ).strftime("%I:%M %p").lstrip("0")
                                  if ca and isinstance(ca, datetime) else "N/A")
                    metodo = "Fiado" if order.get('payment_method') == 'fiado' else "Efectivo"
                    header_txt = (f"**#{idx}** — {order.get('title','Sin título')} "
                                  f"| ${order.get('price',0):,.2f} | {metodo} | {order_time}")
                    with st.expander(header_txt):
                        cust = order.get('customer_name', '')
                        if cust and cust != "Cliente General":
                            st.markdown(f"👤 **Cliente:** {cust}")
                        st.markdown("**Productos:**")
                        for item in order.get('ingredients', []):
                            n = item.get('name', 'N/A')
                            q = item.get('quantity', 0)
                            p = item.get('sale_price', 0.0)
                            c_u = item.get('purchase_price', 0.0)
                            mg  = (p - c_u) * q
                            st.markdown(
                                f"- **{n}** × {q} · Venta: ${p*q:,.2f} "
                                f"· Costo: ${c_u*q:,.2f} · Margen: ${mg:,.2f}")

                st.markdown("---")

                # ── 5. Descarga Excel profesional ─────────────────────
                st.subheader("📥 Descargar Reporte Profesional (Excel)")
                st.info("El Excel incluye **7 hojas**: Portada · Resumen Diario · "
                        "Ventas Detalladas · Productos Hoy · Clientes Fiado · "
                        "Resumen Mensual · Productos del Mes — con colores, "
                        "márgenes de beneficio y análisis de costos.")

                excel_buf = _build_professional_excel(
                    _fecha_legible, _hora_legible, _mes_actual,
                    completed_orders_today, items_sold, df_items_day,
                    total_revenue, total_cash, total_credit,
                    num_transactions, avg_ticket,
                    total_cogs_day, gross_profit_day,
                    completed_orders_month, month_revenue, month_cogs,
                    month_profit, month_margin, month_num_sales,
                    month_items_dict,
                )
                st.download_button(
                    label="⬇️ Descargar Reporte Completo Profesional (Excel)",
                    data=excel_buf,
                    file_name=f"Reporte_SAVA_{today_utc.strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True,
                )

        # ────────────────────────────────────────────────────────────────
        #  TAB 2 — ANÁLISIS MENSUAL
        # ────────────────────────────────────────────────────────────────
        with tab_mes:
            st.subheader(f"📅 Análisis Mensual — {_mes_actual}")

            if not completed_orders_month:
                st.warning(f"No hay ventas completadas en {_mes_actual}.")
            else:
                # ── 1. KPIs mensuales ─────────────────────────────────
                mk1, mk2, mk3, mk4, mk5 = st.columns(5)
                mk1.metric("💰 Ingresos del Mes",    f"${month_revenue:,.2f}")
                mk2.metric("📦 Costo Acumulado",     f"${month_cogs:,.2f}")
                mk3.metric("📈 Beneficio Bruto",     f"${month_profit:,.2f}")
                mk4.metric("📊 Margen del Mes",      f"{month_margin:.1f}%")
                mk5.metric("🧾 Ventas del Mes",      month_num_sales)

                st.markdown("---")

                # ── 2. Gráfico de tendencia ───────────────────────────
                st.subheader("📈 Tendencia de Ventas Diarias del Mes")
                import plotly.express as px  # Lazy import

                daily_data = []
                for o in completed_orders_month:
                    ca = o.get('completed_at')
                    if ca and isinstance(ca, datetime):
                        daily_data.append({
                            'Fecha':    ca.date(),
                            'Ingresos': o.get('price', 0)
                        })

                if daily_data:
                    df_daily = pd.DataFrame(daily_data)
                    df_daily['Fecha'] = pd.to_datetime(df_daily['Fecha'])
                    df_agg = df_daily.groupby('Fecha').agg(
                        Ingresos=('Ingresos', 'sum'),
                        Ventas=('Ingresos', 'count')
                    ).reset_index()

                    fig = px.bar(
                        df_agg, x='Fecha', y='Ingresos',
                        title=f"Ingresos Diarios — {_mes_actual}",
                        labels={'Ingresos': 'Ingresos ($)', 'Fecha': 'Día'},
                        text_auto='$.2s', color='Ingresos',
                        color_continuous_scale='Greens',
                    )
                    fig.update_layout(
                        xaxis_tickformat='%d %b',
                        yaxis_tickprefix='$', yaxis_tickformat=',.0f',
                        showlegend=False,
                    )
                    st.plotly_chart(fig, use_container_width=True)

                st.markdown("---")

                # ── 3. Top productos del mes ──────────────────────────
                st.subheader("🏆 Top Productos del Mes")
                if month_items_dict:
                    df_top = pd.DataFrame(list(month_items_dict.values()))
                    df_top = df_top.sort_values('Ingresos', ascending=False).head(10)
                    st.dataframe(
                        df_top, use_container_width=True, hide_index=True,
                        column_config={
                            "Ingresos": st.column_config.NumberColumn(format="$%.2f")
                        }
                    )

                st.markdown("---")

                # ── 4. Comparativa Hoy vs. Mes ────────────────────────
                st.subheader("⚡ Comparativa: Hoy vs. Mes")
                day_pct      = (total_revenue / month_revenue * 100) if month_revenue > 0 else 0
                month_avg_d  = month_revenue / max(_now_col.day, 1)
                diff_vs_avg  = ((total_revenue - month_avg_d) / month_avg_d * 100) if month_avg_d > 0 else 0

                c1, c2, c3 = st.columns(3)
                c1.metric("📊 Participación del Día", f"{day_pct:.1f}%")
                c2.metric("📈 Promedio Diario del Mes", f"${month_avg_d:,.2f}")
                c3.metric("🔄 Hoy vs. Promedio",
                          f"${total_revenue:,.2f}",
                          delta=f"{diff_vs_avg:+.1f}%")

                st.markdown("---")

                # ── 5. Descarga Excel (mismo botón desde aquí también) ─
                st.subheader("📥 Descargar Reporte Completo")
                excel_buf2 = _build_professional_excel(
                    _fecha_legible, _hora_legible, _mes_actual,
                    completed_orders_today, items_sold, df_items_day,
                    total_revenue, total_cash, total_credit,
                    num_transactions, avg_ticket,
                    total_cogs_day, gross_profit_day,
                    completed_orders_month, month_revenue, month_cogs,
                    month_profit, month_margin, month_num_sales,
                    month_items_dict,
                )
                st.download_button(
                    label="⬇️ Descargar Reporte Completo Profesional (Excel)",
                    data=excel_buf2,
                    file_name=f"Reporte_SAVA_{today_utc.strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True,
                    key="dl_excel_mes",
                )

    except Exception as e:
        st.error(f"Error al generar el reporte diario: {e}")


elif st.session_state.page == "🏢 Acerca de SAVA":
    st.title("Sobre SAVA SOFTWARE")
    st.subheader("Innovación y Tecnología para el Retail del Futuro")

    st.markdown("""
    En **SAVA**, somos pioneros en el desarrollo de soluciones de software que fusionan la inteligencia artificial
    con las necesidades reales del sector retail. Nuestra misión es empoderar a los negocios con herramientas
    poderosas, intuitivas y eficientes que transformen sus operaciones y potencien su crecimiento.

    Creemos que la tecnología debe ser un aliado, no un obstáculo. Por eso, diseñamos **Rapi Tienda Acuarela** pensando
    en la agilidad, la precisión y la facilidad de uso.
    """)

    st.markdown("---")

    st.subheader("Nuestro Equipo Fundador")

    col1_founder, col2_founder = st.columns([1, 3])
    with col1_founder:
        st.image("https://github.com/GIUSEPPESAN21/LOGO-SAVA/blob/main/LOGO%20COLIBRI.png?raw=true", width=200, caption="CEO")
    with col2_founder:
        st.markdown("#### Joseph Javier Sánchez Acuña")
        st.markdown("**CEO - SAVA SOFTWARE FOR ENGINEERING**")
        st.write("""
        Líder visionario con una profunda experiencia en inteligencia artificial y desarrollo de software.
        Joseph es el cerebro detrás de la arquitectura de Rapi Tienda Acuarela, impulsando la innovación
        y asegurando que nuestra tecnología se mantenga a la vanguardia.
        """)
        st.markdown(
            """
            - **LinkedIn:** [joseph-javier-sánchez-acuña](https://www.linkedin.com/in/joseph-javier-sánchez-acuña-150410275)
            - **GitHub:** [GIUSEPPESAN21](https://github.com/GIUSEPPESAN21)
            """
        )
    st.markdown("---")

    st.markdown("##### Cofundadores")

    c1_cof, c2_cof, c3_cof = st.columns(3)
    with c1_cof:
        st.info("**Xammy Alexander Victoria Gonzalez**\n\n*Director Comercial*")
    with c2_cof:
        st.info("**Jaime Eduardo Aragon Campo**\n\n*Director de Operaciones*")
    with c3_cof:
        st.info("**Joseph Javier Sanchez Acuña**\n\n*Director de Proyecto*")
